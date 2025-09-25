import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import os
import pydicom
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import ctypes
import math
import numpy as np
import pickle
from datetime import datetime


### Aug 15 ###
# Compare Measurements feature in progress
# Can open working files saved from different devices
# Better click/drag for measurement end points
# Vertical slider for frame navigation
# Display and export OR
# m key to start measuring
# Window stays selected/active to detect m key press
# Window levelling
# Save and load working file
# Fixed measurement method (projection onto bone line)

try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

class DICOMViewer:
    def __init__(self, root):

        # Main root window
        self.root = root
        self.root.title("Hand DICOM Viewer with Measurements")
        self.current_working_file = None  # Store path of the loaded .dcmstate file
        self.frame_index = 0
        self.num_frames = 1
        self.dicom = None
        self.pixel_data = None
        self.pixel_spacing = [1.0, 1.0]
        self.resize_after_id = None

        # Measurement tools
        self.measurements = {}          # dicts: {frame_index: {'h': (p1, p2), 'H': (p1, p2)}, ...}
        self.points = []                # list: storage for clicked points eg., [(x1, y1), (x2, y2)]
        self.selected_point = None
        self.dragging = False
        self.bone_lines = {}            # dict: {frame_index: (p1, p2)}     - dashed cyan line
        self.bone_slope = {}            # dict: {frame_index: float}        - slope value of bone line
        self.hx2_Hx1 = None             # float: temporary storage of x2 for H measurement alignment with h
        self.measure_step = None

        # Copy measurements to range of frames
        self.control_frame = tk.Frame(self.root)
        self.control_frame.pack(side=tk.RIGHT, fill=tk.Y)

        ## 3-COLUMN MAIN FRAME ##
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill='both', expand=True)
        self.main_frame.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)  # Canvas expands
        self.main_frame.columnconfigure(1, weight=0)  # Slider fixed width
        self.main_frame.columnconfigure(2, weight=0)  # Controls fixed width

        ## Matplotlib figure and canvas
        self.figure, self.ax = plt.subplots()
        self.ax.axis('off')
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.main_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky='nsew')

        ## Frame navigation slider (vertical between canvas and controls)
        self.slider = tk.Scale(self.main_frame, from_=1, to=1, orient=tk.VERTICAL, command=self.slider_moved, showvalue=False)
        self.slider.grid(row=0, column=1, sticky='ns', padx=5)
        self.slider.config(state='disabled')

        ## Control panel (right side)
        self.control_frame = tk.Frame(self.main_frame, width=250)
        self.control_frame.grid(row=0, column=2, sticky='ns')
        self.control_frame.grid_propagate(False)  # Keep fixed width

        # Window level sliders
        # Window Center (Brightness)
        wc_frame = tk.Frame(self.control_frame)
        tk.Label(wc_frame, text="Brightness").pack(side=tk.LEFT)
        self.wc_slider = tk.Scale(wc_frame, from_=0, to=255, length=250, orient=tk.HORIZONTAL, command=self.update_window_center)
        #Window Width (Contrast)
        ww_frame = tk.Frame(self.control_frame)
        tk.Label(ww_frame, text="Contrast").pack(side=tk.LEFT)
        self.ww_slider = tk.Scale(ww_frame, from_=1, to=512, length=250, orient=tk.HORIZONTAL, command=self.update_window_width)

        # File buttons
        file_frame = tk.Frame(self.control_frame)
        file_frame.pack(pady=5, fill='x')
        self.filename_label = tk.Label(file_frame, text="No file loaded", anchor='w')         # File name
        self.filename_label.pack(side=tk.LEFT, padx=(2, 8))
        tk.Button(file_frame, text="Load DICOM File", command=self.load_file).pack(side=tk.RIGHT)   # Load file button

        # Save working file
        save_frame = tk.Frame(self.control_frame)
        save_frame.pack(pady=5, fill='x')
        tk.Button(save_frame, text="Save Working File", command=self.save_working_file).pack(side=tk.LEFT)
        tk.Button(save_frame, text="Load Working File", command=self.load_working_file).pack(side=tk.RIGHT)
        
        # Save annotated images along with measurements
        export_frame = tk.Frame(self.control_frame)
        export_frame.pack(pady=5, fill='x')
        tk.Button(export_frame, text="Export to Excel", command=self.export_measurements).pack(side=tk.LEFT)
        self.save_images_var = tk.BooleanVar()
        self.save_images_checkbox = tk.Checkbutton(export_frame, text="Include images folder", variable=self.save_images_var).pack(side=tk.LEFT)

        # Row of space
        self.space_0 = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.space_0.pack(pady=5, fill='x')

        prevnext_frame = tk.Frame(self.control_frame)
        prevnext_frame.pack(pady=5, fill='x')
        tk.Button(prevnext_frame, text="<", command=self.prev_frame).pack(side=tk.LEFT)
        tk.Button(prevnext_frame, text=">", command=self.next_frame).pack(side=tk.LEFT)
        self.frame_label = tk.Label(prevnext_frame, text="Frame 1 / 1", width=15)
        self.frame_label.pack(side=tk.LEFT, padx=5)

        # Frame navigation: jump to frame
        nav_frame = tk.Frame(self.control_frame)
        nav_frame.pack(pady=5, fill='x')
        tk.Label(nav_frame, text="Jump to frame").pack(side=tk.LEFT)
        self.jump_entry = tk.Entry(nav_frame, width=8)
        self.jump_entry.pack(side=tk.LEFT, padx=5)
        self.jump_entry.bind('<Return>', lambda event: self.jump_to_frame())
        tk.Button(nav_frame, text="Go", command=self.jump_to_frame).pack(side=tk.LEFT)

        # Row of space
        self.space_1 = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.space_1.pack(pady=5, fill='x')

        # Text area for instructions
        self.text_box = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.text_box.pack(pady=3, fill='x')
        # Text area for h and H measurements
        self.results_box = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.results_box.pack(pady=3, fill='x')

        # Measure buttons 
        measure_frame = tk.Frame(self.control_frame)
        measure_frame.pack(pady=5, fill='x')
        tk.Button(measure_frame, text="Measure", command=self.start_measurement_workflow).pack(side=tk.LEFT)
        tk.Button(measure_frame, text="Clear Measurements", command=self.clear_measurements).pack(side=tk.LEFT)
        self.root.bind('m', lambda event: self.start_measurement_workflow()) # Shortcut to start measuring

        # See measured frames (in a new window)
        summary_frame = tk.Frame(self.control_frame)
        summary_frame.pack(pady=5, fill='x')
        tk.Button(summary_frame, text="Label Frames", command=self.show_frames_measured).pack(fill="x")

        # Copy measurements to range
        copy_frame = tk.Frame(self.control_frame)
        copy_frame.pack(pady=5, fill='x')
        tk.Label(copy_frame, text="Copy to frames (eg. 1-3)").pack(side=tk.LEFT)
        self.range_entry = tk.Entry(copy_frame, width=8)
        self.range_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(copy_frame, text="Copy", command=self.copy_measurements_to_range).pack(side=tk.LEFT)

        # Row of space
        self.space_2 = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.space_2.pack(pady=5, fill='x')

        # Window levelling
        self.window_center = 128
        self.window_width = 256
        self.original_window_center = None
        self.original_window_width = None

        # Center (Brightness)
        self.wc_slider.set(self.window_center)
        self.wc_slider.pack(side=tk.RIGHT, padx=5) 
        wc_frame.pack(pady=5, fill='x')
        # Width (Contrast)
        self.ww_slider.set(self.window_width)
        self.ww_slider.pack(side=tk.RIGHT, padx=5)
        ww_frame.pack(pady=5, fill='x') 

        # Reset window level
        wReset_frame = tk.Frame(self.control_frame)
        wReset_frame.pack(pady=5, fill='x')
        tk.Button(wReset_frame, text="Reset Window Level", command=self.reset_window_level).pack(side=tk.LEFT)

        # Row of space
        self.space_3 = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.space_3.pack(pady=5, fill='x')

        # Zoom in slider
        self.zoom_level = 0         # Zoom in from 0% to 100%
        self.pan_offset = [0, 0]    # [x_offset, y_offset]
        self.is_panning = False
        self.last_pan_xy = None
        zoom_frame = tk.Frame(self.control_frame)
        zoom_frame.pack(pady=5, fill='x')
        tk.Label(zoom_frame, text="Zoom").pack(side=tk.LEFT)
        self.zoom_slider = tk.Scale(zoom_frame, from_=0, to=100, orient=tk.HORIZONTAL, showvalue=False, command=self.on_zoom_change)
        self.zoom_slider.pack(side=tk.LEFT, padx=5, fill='x', expand=True) 
        # Reset zoom
        tk.Button(zoom_frame, text="Reset View", command=self.reset_zoom).pack(side=tk.LEFT)

        # Row of space
        self.space_4 = tk.Label(self.control_frame, text="", anchor='w', justify='left')
        self.space_4.pack(pady=5, fill='x')

        # Compare Measurements button
        tk.Button(self.control_frame, text="Compare Measurements", command=self.compare_measurements).pack(fill='x')

        # Connect Matplotlib events on figure (DICOM image) to call handler methods
        self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.canvas.mpl_connect("button_press_event", lambda event: self.canvas.get_tk_widget().focus_set())
        # Connect Tkinter-level events to handler functions
        self.root.bind('<Left>', self.on_left_key)
        self.root.bind('<Right>', self.on_right_key)
        self.root.bind('<Configure>', self.on_resize)
        # Bind mouse wheel for scrolling frames
        self.canvas.get_tk_widget().bind("<MouseWheel>", self.on_mouse_wheel)      # Windows/Mac
        self.canvas.get_tk_widget().bind("<Button-4>", self.on_mouse_wheel)        # Linux scroll up
        self.canvas.get_tk_widget().bind("<Button-5>", self.on_mouse_wheel)        # Linux scroll down

        # Set focus to the canvas to stay responsive to keyboard events
        self.root.after(100, self.set_initial_focus)
        self.canvas.mpl_connect("button_press_event", lambda event: self.canvas.get_tk_widget().focus_set())

    def set_initial_focus(self):
        self.canvas.get_tk_widget().focus_set()

    def focus_app_window(self):
        self.root.lift()
        self.root.focus_force()

    def load_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("DICOM files", "*.dcm"), ("All files", "*.*")])
        if not filepath:
            return
        try:
            data_set = pydicom.dcmread(filepath)
            if not hasattr(data_set, 'pixel_array'):
                print("DICOM has no pixel data.")
                return

            self.dicom = data_set
            self.pixel_data = data_set.pixel_array
            self.num_frames = self.pixel_data.shape[0] if self.pixel_data.ndim == 3 else 1
            self.pixel_spacing = [float(sp) for sp in data_set.PixelSpacing] if hasattr(data_set, 'PixelSpacing') else [1.0, 1.0]

            self.frame_index = 0
            self.measurements.clear()
            self.bone_lines.clear()
            self.bone_slope.clear()
            self.points.clear()
            self.selected_point = None
            self.dragging = False
            self.pan_offset = [0, 0]
            self.zoom_level = 0
            self.zoom_slider.set(0)
            self.slider.config(to=self.num_frames, state='normal')
            self.slider.set(1)
            self.initialize_window_level_from_pixel_data()
            self.current_working_file = None  # Clear any working file info

            filename = os.path.basename(filepath)
            self.filename_label.config(text=f"File: {filename}")
            self.show_frame()

        except Exception as e:
            print(f"Error loading DICOM: {e}")
        self.focus_app_window()

    def show_frame(self):

        if not self.dicom:                          # If there is no DICOM data loaded:
            self.slider.config(state='disabled')    # hide the slider
            self.ax.clear()                         # make sure the axes are cleared
            self.ax.axis('off')                     # turn off axes
            self.canvas.draw_idle()                 # update the canvas
            return

        # We are mostly working with multi frame DICOMs (3D array), otherwise a single frame is 2D array
        frame = self.pixel_data[self.frame_index] if self.pixel_data.ndim == 3 else self.pixel_data
        # Apply window level
        frame = self.apply_window_level(frame)
        self.ax.clear()
        self.ax.axis('off')

        # frame.shape[1] is width, frame.shape[0] is height
        self.ax.imshow(frame, cmap='gray', aspect='equal', extent=[0, frame.shape[1], frame.shape[0], 0])
        self.figure.subplots_adjust(left=0, right=1, top=1, bottom=0) # remove padding around the image

        self.ax.set_xlim(0, frame.shape[1]) # Set limits to show entire image
        self.ax.set_ylim(frame.shape[0], 0) # Flip y axis for correct orientation

        # Zoom
        img_height, img_width = frame.shape             
        zoom_fraction = self.zoom_level / 100.0         
        zoom_width = img_width * (1 - zoom_fraction)
        zoom_height = img_height * (1 - zoom_fraction)
        # Pan-adjusted center of image
        center_x = img_width / 2 + self.pan_offset[0]
        center_y = img_height / 2 + self.pan_offset[1]
        # Visible bounds of zoomed area
        x0 = max(0, center_x - zoom_width / 2)
        x1 = min(img_width, center_x + zoom_width / 2)
        y0 = max(0, center_y - zoom_height / 2)
        y1 = min(img_height, center_y + zoom_height / 2)
        # Set the axes limits to zoomed area
        self.ax.set_xlim(x0, x1)
        self.ax.set_ylim(y1, y0)  # flip y axis

        # Measurements overlays
        dot = 'o'
        dotsize = 0.5
        linesize = 0.8

        # get { 'h': (p1, p2),  'H': (p1, p2) } for the current frame
        frame_measures = self.measurements.get(self.frame_index, {})
        # for each h and H, draw the lines and points
        for key, color, offset_dir in [('h', 'red', -1), ('H', 'yellow', 1)]:
            if key in frame_measures:
                p1, p2 = frame_measures[key]

                # Get bone slope and normal vector
                bone = self.bone_lines.get(self.frame_index)
                if bone:
                    (x1, y1), (x2, y2) = bone
                    dx, dy = x2 - x1, y2 - y1
                    if dx == dy == 0:
                        offset_x, offset_y = 0, 0
                    else:
                        # Perpendicular unit vector to bone line
                        normal_x, normal_y = -dy, dx
                        length = (normal_x ** 2 + normal_y ** 2) ** 0.5
                        normal_x /= length
                        normal_y /= length

                        offset_amount = 8  # You can tweak this value
                        offset_x = normal_x * offset_amount * offset_dir
                        offset_y = normal_y * offset_amount * offset_dir
                else:
                    offset_x, offset_y = 0, 0

                # Offset drawing only
                p1o = (p1[0] + offset_x, p1[1] + offset_y)
                p2o = (p2[0] + offset_x, p2[1] + offset_y)

                self.ax.plot([p1o[0], p2o[0]], [p1o[1], p2o[1]], color=color, linewidth=linesize)
                self.ax.plot(p1o[0], p1o[1], marker=dot, markersize=dotsize, color=color)
                self.ax.plot(p2o[0], p2o[1], marker=dot, markersize=dotsize, color=color)


        # Draw the bone line (cyan dashed) if it was confirmed for this frame
        if self.frame_index in self.bone_lines:
            p1, p2 = self.bone_lines[self.frame_index]
            self.ax.plot([p1[0], p2[0]], [p1[1], p2[1]], linestyle='dashed', linewidth=linesize, color='cyan', alpha=0.4)  # 0 < alpha < 1 
            self.ax.plot(p1[0], p1[1], marker=dot, markersize=dotsize, color='cyan')
            self.ax.plot(p2[0], p2[1], marker=dot, markersize=dotsize, color='cyan')

        # Bone points before confirmation
        if len(self.points) == 1:   # one point clicked
            self.ax.plot(self.points[0][0], self.points[0][1], marker=dot, markersize=dotsize, color='cyan')

        elif len(self.points) == 2: # two points clicked, connect with dashed line
            p1, p2 = self.points
            self.ax.plot([p1[0], p2[0]], [p1[1], p2[1]], linestyle='dashed', linewidth=linesize, color='cyan')
            self.ax.plot(p1[0], p1[1], marker=dot, markersize=dotsize, color='cyan')
            self.ax.plot(p2[0], p2[1], marker=dot, markersize=dotsize, color='cyan')

        # Update frame index label and slider
        self.frame_label.config(text=f"Frame {self.frame_index + 1} / {self.num_frames}")
        self.slider.set(self.frame_index + 1)
        self.canvas.draw_idle()
        self.update_measurement_label()

    def ask_bone_line_confirmation(self):
        confirm = messagebox.askyesno("Confirm", "Confirm bone line?")
        self.focus_app_window()

        if not confirm:
            self.points.clear()
            self.measure_step = 'bone_start'
            self.text_box.config(text="Step 1: Click left side of bone line")
            self.show_frame()
            return

        dx = self.points[1][0] - self.points[0][0]                              # x_2 - x_1
        dy = self.points[1][1] - self.points[0][1]                              # y_2 - y_1
        slope = dy / dx if dx != 0 else float('inf')
        self.bone_lines[self.frame_index] = tuple(self.points)
        self.bone_slope[self.frame_index] = slope

        self.text_box.config(text="Step 3: Click the edge of the epiphysis")
        self.points.clear()                                                     # Clear point storage for h step
        self.measure_step = 'h_start'
        self.show_frame()
        self.focus_app_window()

    def start_measurement_workflow(self):
        self.measure_step = 'bone_start'
        self.points.clear()
        self.text_box.config(text="Step 1: Click left side of bone line")
        self.focus_app_window()

    def on_key_press(self, event):
        if event.char.lower() == 'm':
            self.start_measurement_workflow()

    def on_mouse_press(self, event):
        if not self.dicom or event.inaxes != self.ax:
            return
        
        x, y = event.xdata, event.ydata              # Get mouse click coordinates

        if event.key == 'shift':                     # Check for Shift + drag to pan
            self.is_panning = True
            self.last_pan_xy = (event.x, event.y)
            return

        if self.measure_step == 'bone_start':
            self.points = [(x, y)]                                              # Store 1st mouse click for bone line
            self.text_box.config(text="Step 2: Click right side of bone line")
            self.measure_step = 'bone_end'
            return

        elif self.measure_step == 'bone_end':
            self.points.append((x, y))                                          # Store 2nd mouse click for bone line
            self.show_frame()                                                   # Draws second point and preview line
            self.root.after(10, self.ask_bone_line_confirmation)                # Delay confirmation so the UI can update
            return                                                              # points list is now cleared

        elif self.measure_step == 'h_start':                                    # **** h MEASUREMENT BEGINS **** #
            self.points = [(x, y)]                                              # Store 1st mouse click for h's p1  
            self.text_box.config(text="Step 4: Click the base of the epiphysis")
            self.measure_step = 'h_end'
            return

        elif self.measure_step == 'h_end':
            self.points.append((x, y))                                          # Store 2nd mouse click for h's p2
            p1, p2 = self.points                                                # Define p1 and p2 for h measurement

            bone_p1, bone_p2 = self.bone_lines.get(self.frame_index, (None, None))  # Bone line points of current frame
            if not bone_p1 or not bone_p2:
                return
            
            # Project h's points onto bone line
            proj1 = self.project_point_onto_line(p1, bone_p1, bone_p2)          # h's p1 projected onto bone line     
            proj2 = self.project_point_onto_line(p2, bone_p1, bone_p2)          # h's p2 projected onto bone line   *** synced with H ***

            self.hx2_Hx1 = proj2                                                # Store x coord of h's p2 to be H's p1

            if self.frame_index not in self.measurements:
                self.measurements[self.frame_index] = {}

            self.measurements[self.frame_index]['h'] = (proj1, proj2)           # Store h measurements in current frame's measurements dict 
            self.measurements[self.frame_index]['raw_clicks'] = (p1, p2)        # Only 2 clicks for now, H click comes later

            self.text_box.config(text="Step 5: Click the next joint")           # Prompt next step for H measurement
            self.measure_step = 'H_step'
            self.points.clear()                                                 # Clear point storage for H step
            self.show_frame()
            return

        elif self.measure_step == 'H_step':                                     # **** H MEASUREMENT BEGINS **** #   
            p1 = (x, y)                                                         # Store 3rd mouse click for H's p1
            p2 = self.hx2_Hx1                                                   # Use h's p2 as H's p2
            
            bone_p2, bone_p1 = self.bone_lines.get(self.frame_index, (None, None))  # Bone line points of current frame
            if not bone_p2 or not bone_p1:
                return

            # Project H's points onto bone line
            proj1 = self.project_point_onto_line(p1, bone_p2, bone_p1)          # H's p1 projected onto bone line
            proj2 = self.project_point_onto_line(p2, bone_p2, bone_p1)          # H's p2 projected onto bone line   *** synced with h ***

            self.measurements[self.frame_index]['H'] = (proj1, proj2)           # Store H measurements in current frame's measurements dict

            # Update raw_clicks tuple to include H's first click
            click1, click2 = self.measurements[self.frame_index]['raw_clicks']
            self.measurements[self.frame_index]['raw_clicks'] = (click1, click2, p1)

            self.measure_step = None                                            
            self.show_frame()
            return

        frame_measures = self.measurements.get(self.frame_index, {})    # Current frame's measurements {'h': (p1, p2), 'H': (p1, p2)}

        threshold_line = 15      # bigger threshold for clicking/dragging the line
        threshold_endpoints = 5  # smaller threshold for endpoints

        # Get squared distance between 2 points
        def dist_sq(a, b):
            return (a[0] - b[0])**2 + (a[1] - b[1])**2

        # Instead of looping directly, collect possible hits and pick the closest
        possible_hits = []
        for key, (p1, p2) in frame_measures.items():
            offset_dir = -1 if key == 'h' else 1
            bone = self.bone_lines.get(self.frame_index)
            if bone:
                (x1, y1), (x2, y2) = bone
                dx, dy = x2 - x1, y2 - y1
                if dx == dy == 0:
                    offset_x = offset_y = 0
                else:
                    normal_x, normal_y = -dy, dx
                    length = (normal_x ** 2 + normal_y ** 2) ** 0.5
                    normal_x /= length
                    normal_y /= length
                    offset_amount = 5
                    offset_x = normal_x * offset_amount * offset_dir
                    offset_y = normal_y * offset_amount * offset_dir
            else:
                offset_x = offset_y = 0

            p1_vis = (p1[0] + offset_x, p1[1] + offset_y)
            p2_vis = (p2[0] + offset_x, p2[1] + offset_y)

            # Check endpoints
            if dist_sq((x, y), p1_vis) < threshold_endpoints**2:
                possible_hits.append((math.hypot(x - p1_vis[0], y - p1_vis[1]), key, 'p1'))
            if dist_sq((x, y), p2_vis) < threshold_endpoints**2:
                possible_hits.append((math.hypot(x - p2_vis[0], y - p2_vis[1]), key, 'p2'))

            # Pick the closest hit, but endpoints always win over lines
            if possible_hits:
                type_priority = lambda part: 0 if part in ('p1', 'p2') else 1
                _, key, part = min(possible_hits, key=lambda h: (type_priority(h[2]), h[0]))
                self.selected_point = (key, part)
                self.dragging = True
                if part == 'line':
                    self.drag_offset = (x, y)
                print(f"clicked on {key} {part}")
                return  # <<< prevent second block from running

        for key, (p1, p2) in frame_measures.items():
            # Offset drawing logic for h and H
            offset_dir = -1 if key == 'h' else 1
            bone = self.bone_lines.get(self.frame_index)
            if bone:
                (x1, y1), (x2, y2) = bone
                dx, dy = x2 - x1, y2 - y1
                if dx == dy == 0:
                    offset_x = offset_y = 0
                else:
                    normal_x, normal_y = -dy, dx
                    length = (normal_x ** 2 + normal_y ** 2) ** 0.5
                    normal_x /= length
                    normal_y /= length
                    offset_amount = 5
                    offset_x = normal_x * offset_amount * offset_dir
                    offset_y = normal_y * offset_amount * offset_dir
            else:
                offset_x = offset_y = 0

            # Apply visual offset to p1 and p2
            p1_vis = (p1[0] + offset_x, p1[1] + offset_y)
            p2_vis = (p2[0] + offset_x, p2[1] + offset_y)

            if dist_sq((x, y), p1_vis) < threshold_endpoints**2:
                self.selected_point = (key, 'p1')
                self.dragging = True
                print("clicked on p1")
                return
            if dist_sq((x, y), p2_vis) < threshold_endpoints**2:
                self.selected_point = (key, 'p2')
                self.dragging = True
                print("clicked on p2")
                return
            if self.point_near_line((x, y), p1_vis, p2_vis, threshold_line):
                self.selected_point = (key, 'line')
                self.dragging = True
                self.drag_offset = (x, y)
                print("clicked on line")
                return

    def on_mouse_move(self, event):

        if self.is_panning and self.last_pan_xy:
            dx = event.x - self.last_pan_xy[0]
            dy = event.y - self.last_pan_xy[1]
            self.last_pan_xy = (event.x, event.y)

            img_height, img_width = self.pixel_data[0].shape if self.pixel_data.ndim == 3 else self.pixel_data.shape
            zoom_fraction = self.zoom_level / 100.0
            visible_width = img_width * (1 - zoom_fraction)
            visible_height = img_height * (1 - zoom_fraction)

            # Convert screen pixels to data units
            bbox = self.ax.get_window_extent().transformed(self.figure.dpi_scale_trans.inverted())
            ax_width, ax_height = bbox.width * self.figure.dpi, bbox.height * self.figure.dpi

            if ax_width == 0 or ax_height == 0:
                return

            data_dx = dx * visible_width / ax_width
            data_dy = dy * visible_height / ax_height

            self.pan_offset[0] -= data_dx
            self.pan_offset[1] += data_dy  # Inverted y-axis

            self.show_frame()
            return

        if not self.dicom or not self.dragging or event.inaxes != self.ax or not self.selected_point:
            return
        
        x, y = event.xdata, event.ydata # get mouse position coordinates on DICOM image
        if x is None or y is None:
            return
        
        # DRAG h or H (user may want to adjust end points, maintaining parallel slope)  
        key, part = self.selected_point                      # h or H : p1 or p2 or line
        p1, p2 = self.measurements[self.frame_index][key]    # p1 and p2 of measurement line
        slope = self.bone_slope.get(self.frame_index, 0)     # slope of measurement line
        
        if part == 'p1':
            dy = (p2[0] - x) * slope                                    # maintain slope while adjusting p1
            new_p1 = (x, p2[1] - dy)
            self.measurements[self.frame_index][key] = (new_p1, p2)     # update measurement

        elif part == 'p2':
            dx = x - p2[0]
            dy = y - p2[1]
            new_p2_x = p2[0] + dx
            new_p2_y = p2[1] + dy

            # Maintain slope based on p1
            adjusted_new_p2_y = p1[1] + (new_p2_x - p1[0]) * slope
            self.measurements[self.frame_index][key] = (p1, (new_p2_x, adjusted_new_p2_y))

            # Update the other line 
            other_key = 'H' if key == 'h' else 'h'
            if other_key in self.measurements[self.frame_index]:
                op1, op2 = self.measurements[self.frame_index][other_key]

                # Move op2 by the same delta
                new_op2_x = op2[0] + dx
                new_op2_y = op2[1] + dy

                # Correct op2 to maintain slope from op1
                corrected_op2_y = op1[1] + (new_op2_x - op1[0]) * slope
                self.measurements[self.frame_index][other_key] = (op1, (new_op2_x, corrected_op2_y))

        # DRAG LINE (user may want to adjust the position of line, maintining orthogonality to bone line)
        elif part == 'line':
            slope = self.bone_slope.get(self.frame_index, 0)

            if slope == float('inf'):
                dx = x - self.drag_offset[0]
                dy = 0
            elif slope == 0:
                dx = 0
                dy = y - self.drag_offset[1]
            else:
                # Get perpendicular unit vector to bone line
                perp_dx = 1
                perp_dy = -1 / slope
                norm = math.hypot(perp_dx, perp_dy)
                perp_dx /= norm
                perp_dy /= norm

                # Mouse drag vector
                drag_dx = x - self.drag_offset[0]
                drag_dy = y - self.drag_offset[1]

                # Project drag onto perpendicular vector
                move_amount = drag_dx * perp_dx + drag_dy * perp_dy
                dx = move_amount * perp_dx
                dy = move_amount * perp_dy

            new_p1 = (p1[0] + dx, p1[1] + dy)
            new_p2 = (p2[0] + dx, p2[1] + dy)
            self.measurements[self.frame_index][key] = (new_p1, new_p2)
            self.drag_offset = (x, y)

        self.show_frame()

    def on_mouse_release(self, event):
        if self.is_panning:
            self.is_panning = False
            self.last_pan_xy = None
            return

        self.dragging = False
        self.selected_point = None
        self.show_frame()

    def point_near_line(self, pt, line_start, line_end, threshold):
        x, y = pt
        x1, y1 = line_start
        x2, y2 = line_end
        if (x1 == x2) and (y1 == y2):
            return False
        vx, vy = x2 - x1, y2 - y1
        wx, wy = x - x1, y - y1
        c1 = vx * wx + vy * wy
        if c1 <= 0:
            return (x - x1)**2 + (y - y1)**2 <= threshold**2
        c2 = vx * vx + vy * vy
        if c2 <= c1:
            return (x - x2)**2 + (y - y2)**2 <= threshold**2
        b = c1 / c2
        pbx = x1 + b * vx
        pby = y1 + b * vy
        return (x - pbx)**2 + (y - pby)**2 <= threshold**2

    def calculate_distance(self, p1, p2):
        dx = (p2[0] - p1[0]) * self.pixel_spacing[0]
        dy = (p2[1] - p1[1]) * self.pixel_spacing[1]
        return (dx ** 2 + dy ** 2) ** 0.5
    
    def project_point_onto_line(self, pt, line_p1, line_p2):    # Perpendicularly project pt onto line defined by line_p1 and line_p2

        x0, y0 = pt                     # point to project onto bone line
        x1, y1 = line_p1                # bone line start
        x2, y2 = line_p2                # bone line end
        dx, dy = x2 - x1, y2 - y1       # direction vector of bone line
        if dx == 0 and dy == 0:
            return (x1, y1)
        
        # Calculate the projection of point onto the line (dot product divided by direction vector squared)
        scalar_proj = ((x0 - x1)*dx + (y0 - y1)*dy) / (dx*dx + dy*dy)
        # Projected point coordinates
        return (x1 + scalar_proj*dx, y1 + scalar_proj*dy)

    def update_measurement_label(self):
        frame_measures = self.measurements.get(self.frame_index, {})
        h_dist = H_dist = None
        if 'h' in frame_measures:
            h_dist = self.calculate_distance(*frame_measures['h']) # calculate_distance(p1_h, p2_h)
        if 'H' in frame_measures:
            H_dist = self.calculate_distance(*frame_measures['H']) # calculate_distance(p1_H, p2_H)

        if h_dist and H_dist and self.measure_step is None:
            or_ratio = (h_dist / H_dist) * 100 if H_dist != 0 else 0
            self.text_box.config(text="Measurements complete. Drag to adjust.")
            self.results_box.config(
                text=f"h: {h_dist:.2f} mm        H: {H_dist:.2f} mm        OR: {or_ratio:.1f} %")
        elif self.measure_step is not None:
            self.text_box.config(text=self.text_box.cget("text"))   
        else:
            self.text_box.config(text="")
            self.results_box.config(text="")

    def clear_measurements(self):
        confirm = messagebox.askyesno("Confirm", "Clear all measurements for this frame?")
        self.focus_app_window()
        if not confirm:
            return

        if self.frame_index in self.measurements:
            del self.measurements[self.frame_index]
        if self.frame_index in self.bone_lines:
            del self.bone_lines[self.frame_index]
        if self.frame_index in self.bone_slope:
            del self.bone_slope[self.frame_index]

        self.points.clear()
        self.selected_point = None
        self.dragging = False
        self.measure_step = None
        self.hx2_Hx1 = None
        self.text_box.config(text="")
        self.show_frame()
        self.focus_app_window()

    def calc_dist_from_points(self, pts):
        if not pts or len(pts) != 2:
            return None
        p1, p2 = pts
        return ((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2) ** 0.5
    
    def compare_measurements(self):
        def write_number(ws, row, col, value):
            """Write a float with 2 decimal places, store as number in Excel."""
            if value is None or value == "":
                return
            ws.cell(row=row, column=col, value=round(float(value), 2))
            ws.cell(row=row, column=col).number_format = "0.00"

        # Step 1: Ask user to select two .dcmstate files
        file1 = filedialog.askopenfilename(
            title="Select first .dcmstate file",
            filetypes=[("DICOM Working File", "*.dcmstate")])
        if not file1:
            return

        file2 = filedialog.askopenfilename(
            title="Select second .dcmstate file",
            filetypes=[("DICOM Working File", "*.dcmstate")])
        if not file2:
            return

        try:
            with open(file1, "rb") as f1, open(file2, "rb") as f2:
                data1 = pickle.load(f1)
                data2 = pickle.load(f2)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load working files:\n{e}")
            return

        meas1 = data1.get("measurements", {})
        meas2 = data2.get("measurements", {})

        labels1 = data1.get("frame_joint_labels", {})
        labels2 = data2.get("frame_joint_labels", {})

        pixel_spacing1 = data1.get("pixel_spacing", [1.0, 1.0])
        pixel_spacing2 = data2.get("pixel_spacing", [1.0, 1.0])
        col_spacing = (pixel_spacing1[1] + pixel_spacing2[1]) / 2
        row_spacing = (pixel_spacing1[0] + pixel_spacing2[0]) / 2

        # Prepare new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Click Differences"
        file1_name = os.path.basename(file1)
        file2_name = os.path.basename(file2)

        # Header
        ws.append([
            "Joint",
            f"{file1_name} Frame",
            f"{file2_name} Frame",
            "Click 1 dx (mm)",
            "Click 1 dy (mm)",
            "Click 2 dx (mm)",
            "Click 2 dy (mm)",
            "Click 3 dx (mm)",
            "Click 3 dy (mm)",
            "Click1 Dif (mm)", 
            "Click2 Dif (mm)", 
            "Click3 Dif (mm)"
        ])

        def coord_diff_mm(p1, p2):
            if p1 is None or p2 is None:
                return None, None
            dx = abs(p1[0] - p2[0]) * col_spacing
            dy = abs(p1[1] - p2[1]) * row_spacing
            return dx, dy
        
        def euclidean_dist(dx, dy):
            if dx is None or dy is None:
                return None
            return (dx**2 + dy**2) ** 0.5

        # Compare each joint
        all_joints = set(labels1.values()).union(labels2.values())
        row_idx = 2  # start writing after header

        for joint in all_joints:
            frames1 = [f for f, j in labels1.items() if j == joint]
            frames2 = [f for f, j in labels2.items() if j == joint]
            n_compare = min(len(frames1), len(frames2))

            for i in range(n_compare):
                f1_idx = frames1[i]
                f2_idx = frames2[i]
                f1 = meas1.get(f1_idx, {})
                f2 = meas2.get(f2_idx, {})

                rc1 = f1.get('raw_clicks', (None, None, None))
                rc2 = f2.get('raw_clicks', (None, None, None))

                diffs = [coord_diff_mm(rc1[j], rc2[j]) for j in range(3)]
                dists = [euclidean_dist(dx, dy) for dx, dy in diffs]  # new

                ws.cell(row=row_idx, column=1, value=joint)
                ws.cell(row=row_idx, column=2, value=f1_idx + 1)
                ws.cell(row=row_idx, column=3, value=f2_idx + 1)

                col_idx = 4
                for dx, dy in diffs:
                    write_number(ws, row_idx, col_idx, dx)
                    write_number(ws, row_idx, col_idx + 1, dy)
                    col_idx += 2

                # Distance columns
                for dist in dists:
                    write_number(ws, row_idx, col_idx, dist)
                    col_idx += 1

                row_idx += 1

        for i, col_cells in enumerate(ws.columns, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 12

        # Get base filenames without extensions
        file1_base = os.path.splitext(os.path.basename(file1))[0]
        file2_base = os.path.splitext(os.path.basename(file2))[0]

        # Suggested save name
        suggested_name = f"{file1_base}_vs_{file2_base}.xlsx"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=suggested_name,
            title="Save Differences As")

        if not save_path:
            return

        try:
            wb.save(save_path)
            messagebox.showinfo("Success", f"Measurement differences exported to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")


    def prev_frame(self):
        if self.dicom and self.frame_index > 0:
            self.frame_index -= 1
            self.show_frame()

    def next_frame(self):
        if self.dicom and self.frame_index < self.num_frames - 1:
            self.frame_index += 1
            self.show_frame()

    def slider_moved(self, val):
        frame_num = int(val) - 1
        if frame_num != self.frame_index:
            self.frame_index = frame_num
            self.show_frame()

    def on_mouse_wheel(self, event):
        if not self.dicom or self.num_frames <= 1:
            return
        delta = 0
        # Windows / MacOS
        if hasattr(event, 'delta'):
            delta = event.delta
        # Linux scroll
        elif event.num == 4:  # scroll up
            delta = 120
        elif event.num == 5:  # scroll down
            delta = -120

        if delta > 0:
            if self.frame_index > 0:
                self.frame_index -= 1
        elif delta < 0:
            if self.frame_index < self.num_frames - 1:
                self.frame_index += 1
        else:
            return

        self.slider.set(self.frame_index + 1)
        self.show_frame()

    def jump_to_frame(self):
        val = self.jump_entry.get()
        if val.isdigit():
            frame_num = int(val)
            if 1 <= frame_num <= self.num_frames:
                self.frame_index = frame_num - 1
                self.show_frame()

    def on_left_key(self, event):
        self.prev_frame()

    def on_right_key(self, event):
        self.next_frame()


    def export_measurements(self):
        if not self.measurements:
            messagebox.showinfo("No Data", "There are no measurements to export.")
            self.focus_app_window()
            return

        # Extract base filename without extension
        if self.dicom and hasattr(self.dicom, 'filename'):
            base_filename = os.path.splitext(os.path.basename(self.dicom.filename))[0]
        else:
            base_filename = "measurements"

        suggested_name = base_filename + ".xlsx"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=suggested_name,
            title="Save Measurements As")

        if not save_path:
            return

        # Get pixel spacing (mm per pixel) from DICOM metadata
        row_spacing = col_spacing = 1.0
        if hasattr(self.dicom, "PixelSpacing"):
            try:
                row_spacing = float(self.dicom.PixelSpacing[0])  # mm per pixel vertically
                col_spacing = float(self.dicom.PixelSpacing[1])  # mm per pixel horizontally
            except:
                pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Measurements"

        # Excel Header (first row)
        ws.append([
            "Filename", "Frame", "h (mm)", "H (mm)", "OR (%)",
            "Click 1 x", "Click 1 y",
            "Click 2 x", "Click 2 y",
            "Click 3 x", "Click 3 y"
        ])

        filename = os.path.basename(self.dicom.filename) if self.dicom and hasattr(self.dicom, 'filename') else "Unknown"

        def write_number(ws, row, col, value):
            """Write a float with 2 decimal places to Excel."""
            if value is None or value == "":
                ws.cell(row=row, column=col, value="")
            else:
                ws.cell(row=row, column=col, value=round(float(value), 2))
                ws.cell(row=row, column=col).number_format = "0.00"

        # Write each frame
        for frame_num in sorted(self.measurements.keys()):
            frame_data = self.measurements[frame_num]

            # Optional: keep original h/H distances if present
            h_dist = round(self.calculate_distance(*frame_data['h']), 2) if 'h' in frame_data else None
            H_dist = round(self.calculate_distance(*frame_data['H']), 2) if 'H' in frame_data else None
            or_ratio = round(h_dist / H_dist * 100, 1) if h_dist is not None and H_dist not in (None, 0) else ""

            # Pull raw clicks
            raw_clicks = frame_data.get('raw_clicks', (None, None, None))
            click_coords_mm = []
            for p in raw_clicks:
                if p is not None:
                    click_coords_mm.extend([p[0] * col_spacing, p[1] * row_spacing])
                else:
                    click_coords_mm.extend([None, None])

            # Write row to Excel with 2 decimal places
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value=filename)         # Filename
            ws.cell(row=row_idx, column=2, value=frame_num + 1)   # Frame
            write_number(ws, row_idx, 3, h_dist)                  # h (mm)
            write_number(ws, row_idx, 4, H_dist)                  # H (mm)
            write_number(ws, row_idx, 5, or_ratio)                # OR (%)

            # Clicks 1-3
            col_idx = 6
            for val in click_coords_mm:
                write_number(ws, row_idx, col_idx, val)
                col_idx += 1

        # Save Excel
        try:
            wb.save(save_path)

            # Save images if checkbox is checked
            if self.save_images_var.get():
                image_folder = os.path.splitext(save_path)[0] + "_images"
                self.save_images(image_folder)

            messagebox.showinfo("Success", f"Measurements exported to:\n{save_path}")
            self.focus_app_window()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")

        self.focus_app_window()

    def save_images(self, base_folder):
        # Extract base name of loaded DICOM file (without extension)
        dicom_basename = os.path.splitext(os.path.basename(self.dicom.filename))[0] \
            if self.dicom and hasattr(self.dicom, 'filename') else "DICOM"

        # Save directly to the base folder
        save_folder = base_folder
        os.makedirs(save_folder, exist_ok=True)

        original_index = self.frame_index
        saved_count = 0

        for i in range(self.num_frames):
            frame_measures = self.measurements.get(i, {})
            if not ('h' in frame_measures or 'H' in frame_measures):
                continue

            self.frame_index = i
            frame = self.pixel_data[i] if self.pixel_data.ndim == 3 else self.pixel_data

            fig = Figure(figsize=(6, 6), dpi=150)
            canvas = FigureCanvas(fig)
            ax = fig.add_subplot(111)
            ax.imshow(frame, cmap='gray', aspect='equal', extent=[0, frame.shape[1], frame.shape[0], 0])
            ax.axis('off')

            if i in self.bone_lines:
                p1, p2 = self.bone_lines[i]
                ax.plot([p1[0], p2[0]], [p1[1], p2[1]], linestyle='dashed', color='cyan', linewidth=0.8)

            for key, color, offset_dir in [('h', 'red', -1), ('H', 'yellow', 1)]:
                if key in frame_measures:
                    p1, p2 = frame_measures[key]
                    bone = self.bone_lines.get(i)
                    if bone:
                        (x1, y1), (x2, y2) = bone
                        dx, dy = x2 - x1, y2 - y1
                        if dx == dy == 0:
                            offset_x = offset_y = 0
                        else:
                            normal_x, normal_y = -dy, dx
                            length = (normal_x ** 2 + normal_y ** 2) ** 0.5
                            normal_x /= length
                            normal_y /= length
                            offset_amount = 5
                            offset_x = normal_x * offset_amount * offset_dir
                            offset_y = normal_y * offset_amount * offset_dir
                    else:
                        offset_x = offset_y = 0

                    p1o = (p1[0] + offset_x, p1[1] + offset_y)
                    p2o = (p2[0] + offset_x, p2[1] + offset_y)

                    ax.plot([p1o[0], p2o[0]], [p1o[1], p2o[1]], color=color, linewidth=0.8)
                    ax.plot(p1o[0], p1o[1], marker='o', markersize=0.5, color=color)
                    ax.plot(p2o[0], p2o[1], marker='o', markersize=0.5, color=color)

            out_path = os.path.join(base_folder, f"frame_{i + 1:03d}.png")
            fig.savefig(out_path, bbox_inches='tight', pad_inches=0)
            plt.close(fig)
            saved_count += 1

        self.frame_index = original_index
        self.show_frame()
        self.focus_app_window()


    def copy_measurements_to_range(self):
        if self.frame_index not in self.measurements:
            messagebox.showinfo("No Data", "No measurements found on current frame.")
            self.focus_app_window()
            return

        range_str = self.range_entry.get().strip()
        if not range_str:
            return

        try:
            start_str, end_str = range_str.split('-')
            start = int(start_str) - 1  # Convert to 0-based index
            end = int(end_str) - 1
        except ValueError:
            messagebox.showerror("Invalid Input", "Enter range as start-end (e.g. 5-12)")
            self.focus_app_window()
            return

        if start > end or start < 0 or end >= self.num_frames:
            messagebox.showerror("Out of Range", f"Enter a valid range between 1 and {self.num_frames}")
            return

        source_frame = self.frame_index
        source_measures = {}
        if source_frame in self.measurements:
            if 'h' in self.measurements[source_frame]:
                source_measures['h'] = self.measurements[source_frame]['h'][:]
            if 'H' in self.measurements[source_frame]:
                source_measures['H'] = self.measurements[source_frame]['H'][:]

        source_bone = self.bone_lines.get(source_frame)
        source_slope = self.bone_slope.get(source_frame)

        for i in range(start, end + 1):
            self.measurements[i] = {}
            if 'h' in source_measures:
                self.measurements[i]['h'] = source_measures['h'][:]
            if 'H' in source_measures:
                self.measurements[i]['H'] = source_measures['H'][:]
            if source_bone:
                self.bone_lines[i] = source_bone[:]
            if source_slope is not None:
                self.bone_slope[i] = source_slope
                
        messagebox.showinfo("Success", f"Measurements copied to frames {start+1} to {end+1}.")
        self.focus_app_window()


    def on_resize(self, event):
        if self.dicom is None:
            return
        if self.resize_after_id is not None:
            self.root.after_cancel(self.resize_after_id)
        self.resize_after_id = self.root.after(200, self.show_frame)

    def on_zoom_change(self, val):
        self.zoom_level = int(val)
        self.show_frame()

    def reset_zoom(self):
        self.zoom_level = 0
        self.pan_offset = [0, 0]
        self.zoom_slider.set(0)
        self.show_frame()

    # WINDOW LEVELLING
    def apply_window_level(self, frame):
        wc = self.window_center
        ww = self.window_width
        frame = frame.astype(np.float32)

        # Prevent divide by zero
        if ww < 1:
            ww = 1

        lower = wc - ww / 2
        upper = wc + ww / 2

        frame = np.clip(frame, lower, upper)
        frame = (frame - lower) / (upper - lower) * 255.0
        return frame.astype(np.uint8)
    
    def update_window_center(self, val):
        self.window_center = int(val)
        self.show_frame()

    def update_window_width(self, val):
        self.window_width = int(val)
        self.show_frame()

    def reset_window_level(self):
        if self.original_window_center is not None and self.original_window_width is not None:
            self.window_center = self.original_window_center
            self.window_width = self.original_window_width
            self.wc_slider.set(self.window_center)
            self.ww_slider.set(self.window_width)
            self.show_frame()

    def initialize_window_level_from_pixel_data(self):
        # Estimate initial WC/WW from data
        pixel_min = np.min(self.pixel_data)
        pixel_max = np.max(self.pixel_data)
        self.original_window_center = (pixel_max + pixel_min) // 2
        self.original_window_width = pixel_max - pixel_min
        self.window_center = self.original_window_center
        self.window_width = self.original_window_width

        # Set slider values accordingly
        self.wc_slider.config(from_=int(pixel_min), to=int(pixel_max))
        self.ww_slider.config(from_=1, to=int(pixel_max - pixel_min))
        self.wc_slider.set(self.window_center)
        self.ww_slider.set(self.window_width)


    def save_working_file(self):
        if not self.dicom:
            messagebox.showinfo("No DICOM", "Load a DICOM file first.")
            self.focus_app_window()
            return

        dicom_filename_only = os.path.basename(self.dicom.filename) if getattr(self.dicom, "filename", None) else None

        dicom_name = os.path.splitext(dicom_filename_only)[0] if dicom_filename_only else "DICOM"
        date_str = datetime.now().strftime("%m-%d-%y")
        suggested_name = f"{dicom_name}_{date_str}.dcmstate"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".dcmstate",
            filetypes=[("DICOM Working File", "*.dcmstate")],
            title="Save Working File",
            initialfile=suggested_name
        )

        if not save_path:
            return

        try:
            data = {
                "dicom_path": getattr(self.dicom, "filename", None),  # May be None
                "dicom_filename": dicom_filename_only,                # Just the file name
                "measurements": self.measurements,
                "bone_lines": self.bone_lines,
                "bone_slope": self.bone_slope,
                "frame_index": self.frame_index,
                "zoom_level": self.zoom_level,
                "pan_offset": self.pan_offset,
                "window_center": self.window_center,
                "window_width": self.window_width,
                "original_window_center": self.original_window_center,
                "original_window_width": self.original_window_width,
                "frame_joint_labels": getattr(self, "frame_joint_labels", {})
            }
            with open(save_path, "wb") as f:
                pickle.dump(data, f)
            messagebox.showinfo("Saved", f"Working file saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save working file:\n{e}")
        self.focus_app_window()

    def load_working_file(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("DICOM Working File", "*.dcmstate")],
            title="Load Working File"
        )
        if not filepath:
            return

        self.current_working_file = filepath  # Save loaded working file path

        try:
            with open(filepath, "rb") as f:
                data = pickle.load(f)

            dicom_path = data.get("dicom_path")
            dicom_filename = data.get("dicom_filename")

            # Step 1: Try original saved path
            if dicom_path and os.path.exists(dicom_path):
                pass  # dicom_path is already valid

            # Step 2: Try same folder as .dcmstate
            elif dicom_filename:
                possible_path = os.path.join(os.path.dirname(filepath), dicom_filename)
                if os.path.exists(possible_path):
                    dicom_path = possible_path
                else:
                    dicom_path = None

            # Step 3: Ask user to locate manually
            if not dicom_path or not os.path.exists(dicom_path):
                messagebox.showwarning(
                    "Missing DICOM",
                    f"Please select the corresponding DICOM file in the following window."
                )
                dicom_path = filedialog.askopenfilename(
                    filetypes=[("DICOM files", "*.dcm"), ("All files", "*.*")],
                    title="Locate Original DICOM File for this working file"
                )
                if not dicom_path:
                    self.focus_app_window()
                    return

            # Load DICOM
            self.dicom = pydicom.dcmread(dicom_path)
            self.pixel_data = self.dicom.pixel_array
            self.pixel_spacing = [float(sp) for sp in getattr(self.dicom, 'PixelSpacing', [1.0, 1.0])]
            self.num_frames = self.pixel_data.shape[0] if self.pixel_data.ndim == 3 else 1

            # Restore state
            self.measurements = data.get("measurements", {})
            self.bone_lines = data.get("bone_lines", {})
            self.bone_slope = data.get("bone_slope", {})
            self.frame_index = data.get("frame_index", 0)
            self.zoom_level = data.get("zoom_level", 0)
            self.zoom_slider.set(self.zoom_level)
            self.pan_offset = data.get("pan_offset", [0, 0])
            
            self.window_center = data.get("window_center", self.window_center)
            self.window_width = data.get("window_width", self.window_width)
            self.original_window_center = data.get("original_window_center", self.window_center)
            self.original_window_width = data.get("original_window_width", self.window_width)
            self.wc_slider.set(self.window_center)
            self.ww_slider.set(self.window_width)
            self.initialize_window_level_from_pixel_data()
            self.slider.config(to=self.num_frames, state='normal')

            # Restore joint labels
            self.frame_joint_labels = data.get("frame_joint_labels", {})
            
            # Display working file name (includes date)
            working_filename = os.path.basename(filepath)
            self.filename_label.config(text=f"File: {working_filename}")

            self.show_frame()

            # ---- Print measurements neatly ----
            print("\n=== Loaded Measurements ===")
            if not self.measurements:
                print("No measurements found.")
            else:
                for frame, meas in sorted(self.measurements.items()):
                    print(f"Frame {frame}:")
                    for key, value in meas.items():
                        print(f"  {key}: {value}")
            print("===========================\n")
            # -----------------------------------

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load working file:\n{e}")
        self.focus_app_window()

    # Open a window listing measured frames with editable joint names.
    def show_frames_measured(self):

        joint_names = ["PD4", "PD3", "PD2", "PD5", "PM5", "PM4", "PM3", "PM2",
                       "PP5", "MC5", "PP4", "MC4", "PP3", "MC3", "PP2", "MC2", "PD1"]

        # Ensure we have a dict to store labels
        if not hasattr(self, "frame_joint_labels"):
            self.frame_joint_labels = {}

        win = tk.Toplevel(self.root)
        win.title("Frames Measured")
        win.geometry("600x700")

        # Treeview setup
        columns = ("Joint", "Frame", "h (mm)", "H (mm)", "OR (%)")
        tree = ttk.Treeview(win, columns=columns, show="headings", height=15)
        tree.pack(fill=tk.BOTH, expand=True)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=100)

        # Populate rows from measurements with alternating batch colors
        sorted_frames = sorted(self.measurements.keys())

        batch_index = 0  # 0 = white, 1 = grey
        last_frame = None

        for frame_num in sorted_frames:
            frame_data = self.measurements[frame_num]
            h_val = round(self.calculate_distance(*frame_data['h']), 2) if 'h' in frame_data else None
            H_val = round(self.calculate_distance(*frame_data['H']), 2) if 'H' in frame_data else None
            or_ratio = round(h_val / H_val * 100, 1) if h_val and H_val else None

            current_label = self.frame_joint_labels.get(frame_num, "")

            # Check if this frame is consecutive
            if last_frame is not None and frame_num != last_frame + 1:
                batch_index = 1 - batch_index  # switch color for new batch

            tag_name = "grey" if batch_index else "white"

            tree.insert("", "end", values=(
                current_label,
                frame_num + 1,
                h_val if h_val is not None else "",
                H_val if H_val is not None else "",
                or_ratio if or_ratio is not None else ""
            ), tags=(tag_name,))

            last_frame = frame_num

        # Configure the colors
        tree.tag_configure("grey", background="#f0f0f0")  # light grey
        tree.tag_configure("white", background="#ffffff")  # white

        # Combobox for editing Joint Name
        combo = ttk.Combobox(win, values=joint_names, state="readonly", height=len(joint_names))
        combo.focus_set()  # ensures arrow is drawn immediately

        def on_click(event):
            # Identify selected row/column
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return
            col = tree.identify_column(event.x)
            if col != "#1":  # Joint column
                return
            row_id = tree.identify_row(event.y)
            if not row_id:
                return

            bbox = tree.bbox(row_id, col)
            if not bbox:
                return

            # Remove any existing selection and hide combo
            tree.selection_remove(tree.selection())
            combo.place_forget()

            # Place combobox
            x, y, w, h = bbox
            combo.place(in_=tree, x=x, y=y, width=w, height=h)
            current_value = tree.set(row_id, "Joint")
            combo.set(current_value)
            combo.focus_set()  # force focus to combo

            # Immediately open dropdown
            combo.event_generate('<Button-1>')  # simulates click on arrow

            # Callback when user selects a new joint
            def on_select(event=None):
                new_val = combo.get()
                tree.set(row_id, "Joint", new_val)
                frame_val = int(tree.set(row_id, "Frame"))
                self.frame_joint_labels[frame_val - 1] = new_val
                combo.place_forget()
                tree.focus_set()  # return focus to tree for next click

            combo.bind("<<ComboboxSelected>>", on_select)
            combo.bind("<FocusOut>", lambda e: (combo.place_forget(), tree.focus_set()))
            
        tree.bind("<Button-1>", on_click)


if __name__ == "__main__":
    root = tk.Tk()
    viewer = DICOMViewer(root)
    root.geometry("1200x800")
    root.mainloop()
